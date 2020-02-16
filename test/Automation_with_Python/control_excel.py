"""
作者：陈笔
时间：2020年2月16日
说明：openpycl只能操控xlsx不能操控xls。
"""
import openpyxl as xl
from openpyxl.chart import BarChart, Reference, Series

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    # print(cell.value)
    # print(xl.load_workbook('transaction.xlsx')['Sheet1'].cell(1, 1).value)  # 用一行代码代替以上四行代码的写法
    sheet.cell(1, 4).value = "corrected price"
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                      min_row=1,
                      max_row=sheet.max_row,
                      min_col=3,
                      max_col=3,
                      )
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "E2")
    wb.save(filename)

process_workbook("transaction.xlsx")